import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ============================================================================
# CONFIGURATION - Update these values with your FreshService credentials
# ============================================================================
FRESHSERVICE_DOMAIN = "hummingbirdhealthcare"  # e.g., "hummingbirdhealthcare"
FRESHSERVICE_API_KEY = "your_api_key_here"    # Replace with your actual API key
BASE_URL = f"https://{FRESHSERVICE_DOMAIN}.freshservice.com/api/v2"

# ============================================================================
# Constants
# ============================================================================
TICKETS_PER_PAGE = 100
SEARCH_FILTER = "Request SaaS Application for Experimentation"
OUTPUT_DIR = "reports"

# Excel column headers and their corresponding FreshService field paths
COLUMN_MAPPING = [
    ("Ticket ID", "id"),
    ("Created Date", "created_at"),
    ("Requester Name", "custom_fields.please_select_your_name"),
    ("Application Name", "custom_fields.application_name"),
    ("Application Website", "custom_fields.application_website"),
    ("Trial?", "custom_fields.trial"),
    ("Monthly Cost", "custom_fields.how_much_does_the_application_cost_per_month"),
    ("Has Funding?", "custom_fields.do_you_have_funding_for_this_application"),
    ("Duration Needed", "custom_fields.how_long_would_you_need_this_application"),
    ("Start Date", "custom_fields.start_date"),
    ("Experiment End Date", "custom_fields.experiment_end_date"),
    ("Internet/Cloud Required?", "custom_fields.does_the_application_require_a_connection_to_the_internet_or_to_the_cloud"),
    ("PII/PHI Involved?", "custom_fields.will_you_be_entering_pii_or_phi_into_the_application"),
    ("Problem Description", "custom_fields.problem_description"),
    ("Size/Impact", "custom_fields.size_description"),
    ("Success Description", "custom_fields.success_description"),
    ("Prior Solutions Explored", "custom_fields.prior_solution_exploration"),
]


def get_nested_value(obj, path):
    """
    Safely retrieve a nested value from a dictionary using dot notation.
    
    Args:
        obj: Dictionary to search
        path: Dot-separated path (e.g., "custom_fields.application_name")
    
    Returns:
        The value if found, otherwise None
    """
    keys = path.split(".")
    value = obj
    for key in keys:
        if isinstance(value, dict):
            value = value.get(key)
        else:
            return None
    return value


def fetch_all_tickets():
    """
    Fetch all tickets from FreshService API, paginating through all pages.
    Filter for tickets matching SEARCH_FILTER in the subject.
    
    Returns:
        List of ticket dictionaries
    """
    all_tickets = []
    page = 1
    
    while True:
        print(f"Fetching page {page}...")
        
        url = f"{BASE_URL}/tickets"
        params = {
            "per_page": TICKETS_PER_PAGE,
            "page": page
        }
        
        try:
            response = requests.get(
                url,
                params=params,
                auth=(FRESHSERVICE_API_KEY, "X")  # FreshService uses API key as username, "X" as password
            )
            response.raise_for_status()
            
            data = response.json()
            tickets = data.get("tickets", [])
            
            if not tickets:
                # No more tickets on this page
                break
            
            # Filter for SaaS request tickets
            for ticket in tickets:
                subject = ticket.get("subject", "")
                if SEARCH_FILTER in subject:
                    all_tickets.append(ticket)
            
            page += 1
            
        except requests.exceptions.RequestException as e:
            print(f"Error fetching page {page}: {e}")
            break
    
    print(f"Found {len(all_tickets)} matching tickets")
    return all_tickets


def format_date(date_string):
    """
    Format ISO date string to a readable format (YYYY-MM-DD).
    
    Args:
        date_string: ISO format date string or None
    
    Returns:
        Formatted date string or empty string if None
    """
    if not date_string:
        return ""
    
    try:
        # Parse ISO format date
        date_obj = datetime.fromisoformat(date_string.replace("Z", "+00:00"))
        return date_obj.strftime("%Y-%m-%d")
    except (ValueError, AttributeError):
        return str(date_string)


def create_excel_report(tickets):
    """
    Create a formatted Excel workbook with ticket data.
    
    Args:
        tickets: List of ticket dictionaries
    
    Returns:
        openpyxl.Workbook object
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SaaS Requests"
    
    # Add header row
    headers = [col_name for col_name, _ in COLUMN_MAPPING]
    ws.append(headers)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add data rows with alternating shading
    light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    for row_idx, ticket in enumerate(tickets, start=2):
        row_data = []
        
        for col_name, field_path in COLUMN_MAPPING:
            value = get_nested_value(ticket, field_path)
            
            # Format dates
            if "date" in col_name.lower() and value:
                value = format_date(value)
            
            # Handle None/null values
            if value is None:
                value = ""
            
            row_data.append(value)
        
        ws.append(row_data)
        
        # Apply alternating row shading
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = light_fill
        
        # Center align most cells
        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Auto-size columns
    for idx, (col_name, _) in enumerate(COLUMN_MAPPING, start=1):
        col_letter = get_column_letter(idx)
        max_length = len(col_name)
        
        # Check data rows for width
        for cell in ws[col_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Set header row height
    ws.row_dimensions[1].height = 30
    
    return wb


def save_report(wb):
    """
    Save the workbook to a timestamped Excel file.
    
    Args:
        wb: openpyxl.Workbook object
    
    Returns:
        Path to the saved file
    """
    # Create reports directory if it doesn't exist
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    # Generate timestamped filename
    timestamp = datetime.now().strftime("%Y-%m-%d")
    filename = f"saas_requests_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    wb.save(filepath)
    print(f"Report saved to: {filepath}")
    
    return filepath


def main():
    """Main execution function."""
    print("=" * 80)
    print("FreshService SaaS Request Report Generator")
    print("=" * 80)
    print(f"Domain: {FRESHSERVICE_DOMAIN}")
    print(f"Fetching tickets with subject containing: '{SEARCH_FILTER}'")
    print()  
    # Fetch tickets
    tickets = fetch_all_tickets()
    
    if not tickets:
        print("No matching tickets found.")
        return
    
    # Create Excel report
    print("Creating Excel report...")
    wb = create_excel_report(tickets)
    
    # Save report
    filepath = save_report(wb)
    
    print()  
    print("=" * 80)
    print(f"Success! Report generated with {len(tickets)} tickets.")
    print("=" * 80)


if __name__ == "__main__":
    main()